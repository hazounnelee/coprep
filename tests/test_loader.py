"""Tests for the new consolidated data loader."""
import os

import numpy as np
import pandas as pd
import pytest

from src.data.loader import (
    MATERIAL_PROPS,
    _load_handrecorded,
    _load_integrated,
    _load_materials,
    _load_melt_metal,
    _load_melt_ord,
    _load_react_init,
    _load_react_step,
    get_alldata,
)


# ---------------------------------------------------------------------------
# helpers to create temporary Excel files matching the new folder structure
# ---------------------------------------------------------------------------

def _write_xlsx(path: str, df: pd.DataFrame, **kwargs):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    df.to_excel(path, index=False, engine="openpyxl", **kwargs)


def _make_integrated_file(folder: str, line: int, rows: int = 5):
    """Create a 통합일지 Excel file for a given line number."""
    d = os.path.join(folder, "통합일지")
    os.makedirs(d, exist_ok=True)
    fname = f"복사본 2024년도 {line}라인-전구체2단계 통합일지.xlsx"
    df = pd.DataFrame({
        "No.": range(1, rows + 1),
        "포장 Lot No": [f"PKG-{i:04d}-01" for i in range(1, rows + 1)],
        "값": np.random.randn(rows),
    })
    _write_xlsx(os.path.join(d, fname), df)
    return df


def _make_handrecorded_files(folder: str, line: int, file_count: int = 2, rows: int = 5):
    """Create 수기운전일지 Excel files under a line subfolder."""
    d = os.path.join(folder, "수기운전일지", f"{line}라인")
    os.makedirs(d, exist_ok=True)
    dfs = []
    for i in range(file_count):
        df = pd.DataFrame({
            "시간": range(1, rows + 1),
            "온도": np.random.randn(rows),
        })
        _write_xlsx(os.path.join(d, f"data_{i}.xlsx"), df)
        dfs.append(df)
    return dfs


def _make_melt_metal(folder: str, rows: int = 5):
    """Create MELT_WRK_ORD_METAL.xlsx with header=1 (row 0 is junk)."""
    d = os.path.join(folder, "용해작업실적")
    os.makedirs(d, exist_ok=True)
    path = os.path.join(d, "MELT_WRK_ORD_METAL.xlsx")
    # Row 0 = junk header "1.화면항목", row 1 = real header
    junk_row = pd.DataFrame([["1.화면항목"] + [""] * 2])
    real = pd.DataFrame({
        "생산Lot번호": [f"LOT-{i}" for i in range(rows)],
        "설비명": ["반응기A"] * rows,
        "작업일시": pd.date_range("2024-01-01", periods=rows),
    })
    # Write with openpyxl directly to control row 0
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["1.화면항목", "", ""])
    ws.append(list(real.columns))
    for _, row in real.iterrows():
        ws.append(list(row))
    wb.save(path)
    return real


def _make_melt_ord(folder: str, rows: int = 5):
    """Create MELT_WRK_ORD.xlsx with header=1."""
    d = os.path.join(folder, "용해작업실적")
    os.makedirs(d, exist_ok=True)
    path = os.path.join(d, "MELT_WRK_ORD.xlsx")
    real = pd.DataFrame({
        "원료명": ["황산니켈"] * rows,
        "LOT NO": [f"NISO4-{i:03d}" for i in range(rows)],
        "투입중량": np.random.uniform(100, 500, rows),
    })
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["1.화면항목", "", ""])
    ws.append(list(real.columns))
    for _, row in real.iterrows():
        ws.append(list(row))
    wb.save(path)
    return real


def _make_react_init(folder: str, rows: int = 5):
    """Create 반응투입_LOT_INIT.xlsx with a single header row."""
    d = os.path.join(folder, "반응투입스케줄")
    os.makedirs(d, exist_ok=True)
    path = os.path.join(d, "반응투입_LOT_INIT.xlsx")
    df = pd.DataFrame({
        "생산LOT번호": [f"LOT-{i}" for i in range(rows)],
        "초기시작일시": pd.date_range("2024-01-01", periods=rows),
        "작업시간": [480] * rows,
    })
    _write_xlsx(path, df)
    return df


def _make_react_step(folder: str, rows: int = 10):
    """Create 반응투입_LOT_STEP.xlsx with a single header row."""
    d = os.path.join(folder, "반응투입스케줄")
    os.makedirs(d, exist_ok=True)
    path = os.path.join(d, "반응투입_LOT_STEP.xlsx")
    df = pd.DataFrame({
        "투입STEP": range(1, rows + 1),
        "투입시간": [60.0] * rows,
        "pH": np.random.uniform(11.0, 12.0, rows),
    })
    _write_xlsx(path, df)
    return df


def _make_material_file(folder: str, mat_type: str, suffix: str = "ABCDE",
                         n_lots: int = 3):
    """Create a 3-row-header material file under 원재료/.

    Row 0: main headers
    Row 1: categories
    Row 2: item names
    Row 3+: data
    """
    d = os.path.join(folder, "원재료")
    os.makedirs(d, exist_ok=True)
    fname = f"MATR_{mat_type}_{suffix}_원료검사판정이력_220401_260408.xlsx"
    path = os.path.join(d, fname)

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active

    # Row 0: main headers
    ws.append(["No", "입고일", "입고LOT", "검사유형", "Chemical Composition (C01)",
               "Chemical Composition (C01)", "Chemical Composition (C01)",
               "Initial PH (C03)"])
    # Row 1: categories (repeated for multi-column spans)
    ws.append(["No", "입고일", "입고LOT", "검사유형", "Chemical Composition (C01)",
               "Chemical Composition (C01)", "Chemical Composition (C01)",
               "Initial PH (C03)"])
    # Row 2: item names
    ws.append(["No", "입고일", "입고LOT", "검사항목", "Co", "Mn", "Ni", "pH"])
    # Row 3+: data
    for i in range(n_lots):
        ws.append([i + 1, "2024-01-01", f"{mat_type}-LOT-{i:03d}", "정기",
                   0.1 * (i + 1), 0.2 * (i + 1), 0.3 * (i + 1), 11.0 + i * 0.1])

    wb.save(path)
    return n_lots


def _make_material_file_naoh(folder: str, n_lots: int = 3):
    """Create NAOH material file with NaOH in Chemical Composition."""
    d = os.path.join(folder, "원재료")
    os.makedirs(d, exist_ok=True)
    fname = "MATR_NAOH_ABCDE_원료검사판정이력_220401_260408.xlsx"
    path = os.path.join(d, fname)

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["No", "입고일", "입고LOT", "Chemical Composition (C01)"])
    ws.append(["No", "입고일", "입고LOT", "Chemical Composition (C01)"])
    ws.append(["No", "입고일", "입고LOT", "NaOH"])
    for i in range(n_lots):
        ws.append([i + 1, "2024-01-01", f"NAOH-LOT-{i:03d}", 45.0 + i])
    wb.save(path)
    return n_lots


def _make_material_file_nh4oh(folder: str, n_lots: int = 3):
    """Create NH4OH material file (입고LOT only, no properties used)."""
    d = os.path.join(folder, "원재료")
    os.makedirs(d, exist_ok=True)
    fname = "MATR_NH4OH_ABCDE_원료검사판정이력_220401_260408.xlsx"
    path = os.path.join(d, fname)

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["No", "입고일", "입고LOT"])
    ws.append(["No", "입고일", "입고LOT"])
    ws.append(["No", "입고일", "입고LOT"])
    for i in range(n_lots):
        ws.append([i + 1, "2024-01-01", f"NH4OH-LOT-{i:03d}"])
    wb.save(path)
    return n_lots


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def data_folder(tmp_path):
    """Create a full temporary data folder with all sub-structures."""
    root = str(tmp_path / "데이터폴더")
    _make_integrated_file(root, line=1, rows=5)
    _make_integrated_file(root, line=2, rows=3)
    _make_handrecorded_files(root, line=1, file_count=2, rows=5)
    _make_handrecorded_files(root, line=2, file_count=1, rows=3)
    _make_melt_metal(root, rows=5)
    _make_melt_ord(root, rows=5)
    _make_react_init(root, rows=5)
    _make_react_step(root, rows=10)
    _make_material_file(root, "COSO4")
    _make_material_file(root, "MNSO4")
    _make_material_file(root, "NISO4")
    _make_material_file_naoh(root)
    _make_material_file_nh4oh(root)
    return root


# ---------------------------------------------------------------------------
# Tests: _load_integrated
# ---------------------------------------------------------------------------

class TestLoadIntegrated:
    def test_loads_files_by_line(self, data_folder):
        result = _load_integrated(os.path.join(data_folder, "통합일지"))
        assert "1라인" in result
        assert "2라인" in result
        assert isinstance(result["1라인"], pd.DataFrame)
        assert len(result["1라인"]) == 5
        assert len(result["2라인"]) == 3

    def test_extracts_line_from_filename(self, data_folder):
        result = _load_integrated(os.path.join(data_folder, "통합일지"))
        # Keys should be "1라인", "2라인"
        keys = sorted(result.keys())
        assert keys == ["1라인", "2라인"]

    def test_missing_folder_returns_empty(self, tmp_path):
        result = _load_integrated(str(tmp_path / "nonexistent"))
        assert result == {}


# ---------------------------------------------------------------------------
# Tests: _load_handrecorded
# ---------------------------------------------------------------------------

class TestLoadHandrecorded:
    def test_loads_files_by_line(self, data_folder):
        result = _load_handrecorded(os.path.join(data_folder, "수기운전일지"))
        assert "1라인" in result
        assert "2라인" in result

    def test_returns_list_of_dfs(self, data_folder):
        result = _load_handrecorded(os.path.join(data_folder, "수기운전일지"))
        assert isinstance(result["1라인"], list)
        assert len(result["1라인"]) == 2
        assert len(result["2라인"]) == 1
        assert all(isinstance(df, pd.DataFrame) for df in result["1라인"])

    def test_missing_folder_returns_empty(self, tmp_path):
        result = _load_handrecorded(str(tmp_path / "nonexistent"))
        assert result == {}


# ---------------------------------------------------------------------------
# Tests: _load_melt_metal
# ---------------------------------------------------------------------------

class TestLoadMeltMetal:
    def test_loads_with_header_1(self, data_folder):
        path = os.path.join(data_folder, "용해작업실적", "MELT_WRK_ORD_METAL.xlsx")
        result = _load_melt_metal(path)
        assert isinstance(result, pd.DataFrame)
        assert "생산Lot번호" in result.columns
        assert len(result) == 5

    def test_skips_junk_row(self, data_folder):
        path = os.path.join(data_folder, "용해작업실적", "MELT_WRK_ORD_METAL.xlsx")
        result = _load_melt_metal(path)
        # "1.화면항목" should not appear in the data
        assert "1.화면항목" not in result.values

    def test_missing_file_returns_empty_df(self, tmp_path):
        result = _load_melt_metal(str(tmp_path / "nonexistent.xlsx"))
        assert isinstance(result, pd.DataFrame)
        assert len(result) == 0


# ---------------------------------------------------------------------------
# Tests: _load_melt_ord (same pattern as melt_metal)
# ---------------------------------------------------------------------------

class TestLoadMeltOrd:
    def test_loads_with_header_1(self, data_folder):
        path = os.path.join(data_folder, "용해작업실적", "MELT_WRK_ORD.xlsx")
        result = _load_melt_ord(path)
        assert isinstance(result, pd.DataFrame)
        assert "원료명" in result.columns
        assert len(result) == 5


# ---------------------------------------------------------------------------
# Tests: _load_react_init
# ---------------------------------------------------------------------------

class TestLoadReactInit:
    def test_loads_single_header(self, data_folder):
        path = os.path.join(data_folder, "반응투입스케줄", "반응투입_LOT_INIT.xlsx")
        result = _load_react_init(path)
        assert isinstance(result, pd.DataFrame)
        assert "생산LOT번호" in result.columns
        assert len(result) == 5

    def test_missing_file_returns_empty_df(self, tmp_path):
        result = _load_react_init(str(tmp_path / "nonexistent.xlsx"))
        assert isinstance(result, pd.DataFrame)
        assert len(result) == 0


# ---------------------------------------------------------------------------
# Tests: _load_react_step
# ---------------------------------------------------------------------------

class TestLoadReactStep:
    def test_loads_single_header(self, data_folder):
        path = os.path.join(data_folder, "반응투입스케줄", "반응투입_LOT_STEP.xlsx")
        result = _load_react_step(path)
        assert isinstance(result, pd.DataFrame)
        assert "투입STEP" in result.columns
        assert len(result) == 10


# ---------------------------------------------------------------------------
# Tests: _load_materials
# ---------------------------------------------------------------------------

class TestLoadMaterials:
    def test_loads_all_types(self, data_folder):
        result = _load_materials(os.path.join(data_folder, "원재료"))
        assert "COSO4" in result
        assert "MNSO4" in result
        assert "NISO4" in result
        assert "NAOH" in result
        assert "NH4OH" in result

    def test_coso4_has_입고lot_and_co(self, data_folder):
        result = _load_materials(os.path.join(data_folder, "원재료"))
        df = result["COSO4"]
        cols_lower = [c.lower() for c in df.columns]
        assert "입고lot" in cols_lower
        assert "co" in cols_lower

    def test_mnso4_has_mn_and_ph(self, data_folder):
        result = _load_materials(os.path.join(data_folder, "원재료"))
        df = result["MNSO4"]
        cols_lower = [c.lower() for c in df.columns]
        assert "입고lot" in cols_lower
        assert "mn" in cols_lower
        assert "ph" in cols_lower

    def test_niso4_has_ni(self, data_folder):
        result = _load_materials(os.path.join(data_folder, "원재료"))
        df = result["NISO4"]
        cols_lower = [c.lower() for c in df.columns]
        assert "입고lot" in cols_lower
        assert "ni" in cols_lower

    def test_naoh_has_naoh(self, data_folder):
        result = _load_materials(os.path.join(data_folder, "원재료"))
        df = result["NAOH"]
        cols_lower = [c.lower() for c in df.columns]
        assert "입고lot" in cols_lower
        assert "naoh" in cols_lower

    def test_nh4oh_has_only_입고lot(self, data_folder):
        result = _load_materials(os.path.join(data_folder, "원재료"))
        df = result["NH4OH"]
        cols_lower = [c.lower() for c in df.columns]
        assert "입고lot" in cols_lower

    def test_missing_folder_returns_empty(self, tmp_path):
        result = _load_materials(str(tmp_path / "nonexistent"))
        assert result == {}

    def test_row_count(self, data_folder):
        result = _load_materials(os.path.join(data_folder, "원재료"))
        assert len(result["COSO4"]) == 3
        assert len(result["MNSO4"]) == 3


# ---------------------------------------------------------------------------
# Tests: MATERIAL_PROPS constant
# ---------------------------------------------------------------------------

class TestMaterialProps:
    def test_has_all_types(self):
        assert "COSO4" in MATERIAL_PROPS
        assert "MNSO4" in MATERIAL_PROPS
        assert "NISO4" in MATERIAL_PROPS
        assert "NAOH" in MATERIAL_PROPS
        assert "NH4OH" in MATERIAL_PROPS

    def test_coso4_props(self):
        assert MATERIAL_PROPS["COSO4"] == [("Chemical Composition (C01)", "Co")]

    def test_mnso4_props(self):
        assert MATERIAL_PROPS["MNSO4"] == [
            ("Chemical Composition (C01)", "Mn"),
            ("Initial PH (C03)", "pH"),
        ]

    def test_nh4oh_empty(self):
        assert MATERIAL_PROPS["NH4OH"] == []


# ---------------------------------------------------------------------------
# Tests: get_alldata
# ---------------------------------------------------------------------------

class TestGetAlldata:
    def test_returns_dict_with_all_keys(self, data_folder):
        result = get_alldata(data_folder)
        expected_keys = {
            "통합일지", "수기운전일지", "용해_metal", "용해_ord",
            "반응_init", "반응_step", "원재료",
        }
        assert expected_keys == set(result.keys())

    def test_통합일지_structure(self, data_folder):
        result = get_alldata(data_folder)
        assert isinstance(result["통합일지"], dict)
        assert "1라인" in result["통합일지"]

    def test_수기운전일지_structure(self, data_folder):
        result = get_alldata(data_folder)
        assert isinstance(result["수기운전일지"], dict)
        assert isinstance(result["수기운전일지"]["1라인"], list)

    def test_용해_metal_is_df(self, data_folder):
        result = get_alldata(data_folder)
        assert isinstance(result["용해_metal"], pd.DataFrame)

    def test_용해_ord_is_df(self, data_folder):
        result = get_alldata(data_folder)
        assert isinstance(result["용해_ord"], pd.DataFrame)

    def test_반응_init_is_df(self, data_folder):
        result = get_alldata(data_folder)
        assert isinstance(result["반응_init"], pd.DataFrame)

    def test_반응_step_is_df(self, data_folder):
        result = get_alldata(data_folder)
        assert isinstance(result["반응_step"], pd.DataFrame)

    def test_원재료_structure(self, data_folder):
        result = get_alldata(data_folder)
        assert isinstance(result["원재료"], dict)
        assert "COSO4" in result["원재료"]

    def test_empty_folder(self, tmp_path):
        result = get_alldata(str(tmp_path))
        assert isinstance(result, dict)


# ---------------------------------------------------------------------------
# Tests: debug mode
# ---------------------------------------------------------------------------

class TestDebugMode:
    def test_debug_prints_messages(self, data_folder, capsys):
        get_alldata(data_folder, debug=True)
        captured = capsys.readouterr()
        assert "[DEBUG] loader:" in captured.out

    def test_debug_false_no_output(self, data_folder, capsys):
        get_alldata(data_folder, debug=False)
        captured = capsys.readouterr()
        assert "[DEBUG]" not in captured.out

    def test_debug_missing_folder(self, tmp_path, capsys):
        get_alldata(str(tmp_path), debug=True)
        captured = capsys.readouterr()
        assert "[DEBUG] loader:" in captured.out
